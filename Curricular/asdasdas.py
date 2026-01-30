import pandas as pd
import unicodedata
import re
import os
from pathlib import Path
from openpyxl import load_workbook

# --- IMPORTACIONES LOCALES ---
from prerequisitos_corregido import prerequisitos
from malla_FCA import malla_curricular, optativas_produccion

# --- RUTAS ---
#RUTA_ESTUDIANTES = r"C:\Users\JuanPabloGalindoGóme\Documents\Curricular\Historias_academicas.xlsx"
#RUTA_ESTUDIANTES ="/home/jugalindog/Pasantia/Curricular/Curricular/Historias_academicas.xlsx"
RUTA_ESTUDIANTES = "/home/jugalindog/Pasantia/Curricular/Curricular/Historias_academicas3.xlsx"

# Fuente de donde se extrae % avance y # matrículas
AVANCES_PATH = "/home/jugalindog/Pasantia/Curricular/Curricular/Avances_puntaje.xlsx"

# Archivo de salida
ARCHIVO_SALIDA = "Calculo_y_Proyeccion_Cupos_4.xlsx"

# --- UNIFICAR MALLA + OPTATIVAS ---
malla_completa = {}

# Malla obligatoria
for k, v in malla_curricular.items():
    v_copy = v.copy()
    v_copy["tipo"] = "obligatoria"
    malla_completa[k] = v_copy

# Optativas de producción
for k, v in optativas_produccion.items():
    v_copy = v.copy()
    v_copy["tipo"] = "optativa_produccion"
    malla_completa[k] = v_copy


def normalize_name(name):
    """Convierte a minúsculas, quita espacios y acentos."""
    if not isinstance(name, str):
        return ""
    nfkd_form = unicodedata.normalize("NFKD", name.lower().strip())
    s = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    s = re.sub(r"[-_()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# --- SETS PARA RECONOCER OPTATIVAS DE PRODUCCIÓN ---
OPT_CODIGOS = set(str(v.get("codigo", "")).strip() for v in optativas_produccion.values())
OPT_NOMBRES = set(normalize_name(k) for k in optativas_produccion.keys())


def get_siguiente_semestre(sem_str):
    """Calcula el siguiente periodo académico."""
    if not isinstance(sem_str, str) or "-" not in sem_str:
        return "Proyección Inválida"
    try:
        year_str, period_str = sem_str.split("-")
        year, period = int(year_str), int(period_str[0])
        return f"{year}-2S" if period == 1 else f"{year + 1}-1S"
    except (ValueError, IndexError):
        return "Proyección Inválida"


# --- LÓGICA DE EXENCIONES (PLAN) ---
def obtener_exenciones_por_plan(plan_texto):
    """Devuelve materias a aprobar automáticamente según el plan."""
    if not isinstance(plan_texto, str):
        return []

    CODIGO_MATE_BASICA = "1000001-B"
    CODIGO_LECTO = "1000002-B"
    exenciones = []
    plan_lower = plan_texto.lower()

    # Palabras clave detectadas en tus reportes
    if "no deben nivelar" in plan_lower:
        exenciones.extend([CODIGO_MATE_BASICA, CODIGO_LECTO])
    elif "deben nivelar matemáticas" in plan_lower:
        exenciones.append(CODIGO_LECTO)  # Ya aprobó lecto
    elif ("deben nivelar lecto" in plan_lower) and "no deben" not in plan_lower:
        exenciones.append(CODIGO_MATE_BASICA)  # Ya aprobó mate

    return exenciones


# --- VERIFICACIÓN ROBUSTA DE PRERREQUISITOS ---
def verificar_prerequisitos(historial_estudiante, prerequisitos_asignatura, mode, debug=False):
    """
    Verifica prerrequisitos manejando errores de formato, códigos faltantes y lógica AND/OR.
    """
    # 1. Estandarizar entrada (Lista o Tupla con Operador)
    lista_grupos = []
    operador = "AND"

    # Caso: Tupla ( [[...]], 'OR' )
    if isinstance(prerequisitos_asignatura, tuple) and len(prerequisitos_asignatura) == 2:
        lista_grupos = prerequisitos_asignatura[0]
        operador = prerequisitos_asignatura[1]
    # Caso: Lista antigua [[...]]
    elif isinstance(prerequisitos_asignatura, list):
        # Soporta formato legacy: [[...], [...], 'OR'] o [[...], 'AND']
        if prerequisitos_asignatura and isinstance(prerequisitos_asignatura[-1], str):
            op = prerequisitos_asignatura[-1].strip().upper()
            if op in ("OR", "AND"):
                operador = op
                lista_grupos = prerequisitos_asignatura[:-1]  # quita el operador final
            else:
                operador = "AND"
                lista_grupos = prerequisitos_asignatura
        else:
            operador = "AND"
            lista_grupos = prerequisitos_asignatura
    

    # Si la lista está vacía o el primer elemento es vacío
    if not lista_grupos or (len(lista_grupos) > 0 and not lista_grupos[0]):
        return True

    # 2. Pre-cargar materias aprobadas para búsqueda rápida
    historial_aprobado = historial_estudiante[historial_estudiante["estado"].str.strip() == "Aprobada"]
    if mode == "codigo":
        aprobadas_set = set(historial_aprobado["codigo_asignatura"])
    else:
        aprobadas_set = set(historial_aprobado["asignatura_normalizada"])

    resultados_grupos = []

    for grupo in lista_grupos:
        # Manejo de 'Ninguno'
        if isinstance(grupo, (list, tuple)) and len(grupo) > 0:
            primer_elem = grupo[0]
            if isinstance(primer_elem, (list, tuple)) and len(primer_elem) > 0 and primer_elem[0] == "Ninguno":
                resultados_grupos.append(True)
                continue

        cumple_grupo = True

        # Iterar sobre cada materia dentro del grupo (Lógica interna del grupo suele ser AND)
        for prereq in grupo:
            nombre_req = ""
            codigo_req = None

            # --- BLINDAJE: Extracción segura de datos ---
            try:
                if isinstance(prereq, (list, tuple)):
                    if len(prereq) >= 2:
                        nombre_req, codigo_req = prereq[0], prereq[1]
                    elif len(prereq) == 1:
                        nombre_req = prereq[0]
                else:
                    # Formato desconocido, saltar
                    continue
            except Exception:
                continue

            # Limpieza de tuplas anidadas extrañas (ej: (('Nombre',), 'Cod'))
            if isinstance(nombre_req, tuple):
                nombre_req = nombre_req[0]

            # Verificación
            aprobada = False
            if mode == "codigo":
                if codigo_req and str(codigo_req).strip() in aprobadas_set:
                    aprobada = True
            else:  # mode == 'nombre'
                if normalize_name(nombre_req) in aprobadas_set:
                    aprobada = True

            if not aprobada:
                cumple_grupo = False
                break

        resultados_grupos.append(cumple_grupo)

    # 3. Resultado Final
    if operador == "OR":
        return any(resultados_grupos)
    else:  # AND
        return all(resultados_grupos)


# =========================
#   POST-PROCESO EXCEL
# =========================
def _header_map(ws):
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        hm[str(v).strip().lower()] = c
    return hm


def _ensure_col(ws, header):
    hm = _header_map(ws)
    key = header.strip().lower()
    if key in hm:
        return hm[key]
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = header
    return new_col


def agregar_avance_y_matriculas_a_detalle(
    archivo_salida,
    fuente_avances=AVANCES_PATH,
    hoja_fuente="Resumen Estudiantes",
    hoja_destino="Detalle",
):
    """
    Lee Documento + Porcentaje de Avance + Número de Matrículas desde Avances_puntaje.xlsx
    y agrega/llena columnas 'avance' y 'matriculas' en la hoja Detalle del archivo_salida.
    """
    if not Path(fuente_avances).exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {fuente_avances}")

    # 1) Abrir fuente
    wb_src = load_workbook(fuente_avances, data_only=True)
    if hoja_fuente not in wb_src.sheetnames:
        raise ValueError(f"No existe la hoja '{hoja_fuente}' en {fuente_avances}. Hojas: {wb_src.sheetnames}")
    ws_src = wb_src[hoja_fuente]

    hm_src = _header_map(ws_src)

    col_doc = hm_src.get("documento")
    col_av = hm_src.get("porcentaje de avance") or hm_src.get("avance")
    col_mat = hm_src.get("número de matrículas") or hm_src.get("numero de matriculas") or hm_src.get("matriculas")

    if not col_doc:
        raise ValueError("En Avances_puntaje.xlsx no encontré la columna 'Documento' en la hoja fuente.")
    if not col_av:
        raise ValueError("En Avances_puntaje.xlsx no encontré 'Porcentaje de Avance' o 'avance'.")
    if not col_mat:
        raise ValueError("En Avances_puntaje.xlsx no encontré 'Número de Matrículas' o 'matriculas'.")

    lookup = {}
    for r in range(2, ws_src.max_row + 1):
        doc = ws_src.cell(row=r, column=col_doc).value
        if doc is None:
            continue
        doc_key = str(doc).strip()
        lookup[doc_key] = (ws_src.cell(r, col_av).value, ws_src.cell(r, col_mat).value)

    # 2) Abrir destino y editar Detalle
    wb_out = load_workbook(archivo_salida)
    if hoja_destino not in wb_out.sheetnames:
        raise ValueError(f"No existe la hoja '{hoja_destino}' en {archivo_salida}. Hojas: {wb_out.sheetnames}")
    ws_det = wb_out[hoja_destino]

    hm_det = _header_map(ws_det)
    col_doc_det = hm_det.get("documento")
    if not col_doc_det:
        raise ValueError("En el archivo de salida no encontré la columna 'documento' en Detalle.")

    col_av_det = _ensure_col(ws_det, "avance")
    col_mat_det = _ensure_col(ws_det, "matriculas")

    for r in range(2, ws_det.max_row + 1):
        doc = ws_det.cell(row=r, column=col_doc_det).value
        if doc is None:
            continue
        doc_key = str(doc).strip()
        if doc_key in lookup:
            av, mat = lookup[doc_key]
            ws_det.cell(row=r, column=col_av_det).value = av
            ws_det.cell(row=r, column=col_mat_det).value = mat

    wb_out.save(archivo_salida)


def main():
    mode = ""
    while mode not in ["codigo", "nombre"]:
        mode = input("¿Comparar por 'codigo' o 'nombre'? ").lower().strip()

    print(f"\n--- Iniciando Procesamiento ({mode}) ---")

    if not os.path.exists(RUTA_ESTUDIANTES):
        print(f"❌ Error: No se encuentra {RUTA_ESTUDIANTES}")
        return

    try:
        df_estudiantes = pd.read_excel(RUTA_ESTUDIANTES)

        # Limpieza inicial
        df_estudiantes.columns = [c.lower() for c in df_estudiantes.columns]  # Todo a minúsculas
        df_estudiantes["codigo_asignatura"] = df_estudiantes["codigo_asignatura"].astype(str).str.strip()
        df_estudiantes["estado"] = df_estudiantes["estado"].astype(str).str.strip()
        df_estudiantes["asignatura_normalizada"] = df_estudiantes["asignatura"].apply(normalize_name)

        if "plan" not in df_estudiantes.columns:
            print("⚠️ Advertencia: Columna 'plan' no encontrada. Se asume vacío.")
            df_estudiantes["plan"] = ""

    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
        return

    # Solo contamos obligatorias + el bucket genérico de optativas
    resumen_cupos = {asig: 0 for asig, info in malla_completa.items() if info.get("tipo") != "optativa_produccion"}
    resumen_cupos["Optativa de producción"] = 0
    proyecciones = []

    # Mapa de prerrequisitos normalizado
    prereq_map = {normalize_name(k): v for k, v in prerequisitos.items()}

    estudiantes_unicos = df_estudiantes[["documento", "nombre"]].drop_duplicates().to_dict("records")

    for est in estudiantes_unicos:
        doc = est["documento"]
        nom = est["nombre"]

        # Listas por estudiante
        elegibles_obligatorias = []
        elegibles_optativas = []
        elegibles = []  # <-- esta lista alimenta la hoja 'Detalle'

        # Copia aislada del historial
        historial = df_estudiantes[df_estudiantes["documento"] == doc].copy()

        # 1. INYECTAR EXENCIONES DEL PLAN
        try:
            planes = historial["plan"].dropna()
            plan_txt = str(planes.iloc[0]) if not planes.empty else ""
        except Exception:
            plan_txt = ""

        for cod_ex in obtener_exenciones_por_plan(plan_txt):
            ya_esta = historial[
                (historial["codigo_asignatura"] == cod_ex) & (historial["estado"] == "Aprobada")
            ]
            if ya_esta.empty:
                # Buscar nombre real en la malla (solo obligatorias)
                nom_real = "Nivelación"
                for k, v in malla_curricular.items():
                    if str(v.get("codigo")) == cod_ex:
                        nom_real = k
                        break

                fila = {
                    "documento": doc,
                    "nombre": nom,
                    "codigo_asignatura": cod_ex,
                    "asignatura": nom_real,
                    "asignatura_normalizada": normalize_name(nom_real),
                    "estado": "Aprobada",
                    "semestre_asignatura": "Eximido",
                    "semestre_inicio": "Eximido",
                    "plan": plan_txt,
                }
                historial = pd.concat([historial, pd.DataFrame([fila])], ignore_index=True)

        # 2. DEFINIR SEMESTRE BASE PROYECCIÓN
        ult_sem = historial[historial["semestre_asignatura"] != "Eximido"]["semestre_asignatura"].max()

        if pd.isna(ult_sem):
            sem_inicio_series = historial[historial["semestre_inicio"] != "Eximido"]["semestre_inicio"]
            if not sem_inicio_series.empty:
                sem_base = sem_inicio_series.iloc[0]
            else:
                continue
        else:
            sem_base = get_siguiente_semestre(ult_sem)

        # 3. EVALUAR MATERIAS (obligatorias + optativas producción)
        equivalencias_codigos = {
            "1000012-B": ["1000013-B"],  # Bioestadística se considera vista si tiene Probabilidad
        }

        alias_materias = {
            "bioestadistica fundamental": ["bioestadistica", "probabilidad y estadistica"],
        }

        for asignatura, info in malla_completa.items():
            cod_malla = str(info.get("codigo", "")).strip()
            nombre_malla_norm = normalize_name(asignatura)

            # --- ¿ya la vio? (código, nombre, equivalencias, alias) ---
            ya_vio_la_materia = False

            if cod_malla and (historial["codigo_asignatura"] == cod_malla).any():
                ya_vio_la_materia = True
            elif (historial["asignatura_normalizada"] == nombre_malla_norm).any():
                ya_vio_la_materia = True
            elif cod_malla in equivalencias_codigos:
                codigos_alt = [str(c).strip() for c in equivalencias_codigos[cod_malla]]
                if historial[
                    historial["codigo_asignatura"].isin(codigos_alt)
                    & (historial["estado"].str.strip() == "Aprobada")
                ].shape[0] > 0:
                    ya_vio_la_materia = True
            elif nombre_malla_norm in alias_materias:
                nombres_alt = [normalize_name(n) for n in alias_materias[nombre_malla_norm]]
                if historial["asignatura_normalizada"].isin(nombres_alt).any():
                    ya_vio_la_materia = True

            if ya_vio_la_materia:
                continue

            # --- Si es optativa producción: si ya cursó ese código, no la contamos ---
            if info.get("tipo") == "optativa_produccion":
                if cod_malla and (historial["codigo_asignatura"] == cod_malla).any():
                    continue

            # Prerrequisitos
            reqs = prereq_map.get(normalize_name(asignatura), [])
            cumple = verificar_prerequisitos(historial, reqs, mode)

            if cumple:
                if info.get("tipo") == "optativa_produduccion":
                    # (por si hay typo en tu malla; normalmente es 'optativa_produccion')
                    elegibles_optativas.append({"asignatura": asignatura, "semestre_malla": info.get("semestre", 99)})
                elif info.get("tipo") == "optativa_produccion":
                    elegibles_optativas.append({"asignatura": asignatura, "semestre_malla": info.get("semestre", 99)})
                else:
                    resumen_cupos[asignatura] += 1
                    elegibles_obligatorias.append({"asignatura": asignatura, "semestre_malla": info.get("semestre", 99)})

        # --- REGLA: cada estudiante debe tener mínimo 2 optativas de producción ---
        hist_aprob = historial[historial["estado"].str.strip() == "Aprobada"]

        opt_aprob_por_codigo = hist_aprob["codigo_asignatura"].astype(str).str.strip().isin(OPT_CODIGOS).sum()
        opt_aprob_por_nombre = hist_aprob["asignatura_normalizada"].isin(OPT_NOMBRES).sum()

        opt_aprobadas = int(max(opt_aprob_por_codigo, opt_aprob_por_nombre))
        cupos_necesarios = max(0, 2 - opt_aprobadas)

        # Sumatoria global (bucket) + selección de optativas para Detalle
        optativas_seleccionadas = []
        if cupos_necesarios > 0 and len(elegibles_optativas) > 0:
            elegibles_optativas.sort(key=lambda x: x["semestre_malla"])
            optativas_seleccionadas = elegibles_optativas[: min(cupos_necesarios, len(elegibles_optativas))]
            resumen_cupos["Optativa de producción"] += len(optativas_seleccionadas)

        # 4. PROYECCIÓN INDIVIDUAL (hoja Detalle)
        # Incluye obligatorias elegibles y (si aplica) las optativas seleccionadas para cumplir mínimo 2
        elegibles = elegibles_obligatorias + optativas_seleccionadas

        if elegibles:
            elegibles.sort(key=lambda x: x["semestre_malla"])
            sem_actual = sem_base
            sem_malla_ant = None

            for mat in elegibles:
                if sem_malla_ant is not None and mat["semestre_malla"] > sem_malla_ant:
                    sem_actual = get_siguiente_semestre(sem_actual)

                proyecciones.append(
                    {
                        "documento": doc,
                        "nombre": nom,
                        "asignatura": mat["asignatura"],
                        "semestre_proyectado": sem_actual,
                    }
                )
                sem_malla_ant = mat["semestre_malla"]

    # --- EXPORTAR ---
    try:
        df_res = pd.DataFrame(list(resumen_cupos.items()), columns=["Asignatura", "Estudiantes Aptos"])
        df_res = df_res.sort_values("Estudiantes Aptos", ascending=False)
        df_proy = pd.DataFrame(proyecciones)

        with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
            df_res.to_excel(writer, sheet_name="Resumen", index=False)
            if not df_proy.empty:
                df_proy.to_excel(writer, sheet_name="Detalle", index=False)

        # ✅ Post-proceso: agregar columnas avance y matriculas a Detalle desde Avances_puntaje.xlsx
        agregar_avance_y_matriculas_a_detalle(ARCHIVO_SALIDA)

        print("\n✅ Proceso finalizado con éxito.")
    except Exception as e:
        print(f"\n❌ Error guardando/actualizando Excel: {e}")


if __name__ == "__main__":
    main()
