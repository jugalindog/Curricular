from openpyxl import load_workbook
from pathlib import Path

def find_excel_file(folder=".", pattern="*.xlsx"):
    folder = Path(folder)
    matches = sorted(folder.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No encontré archivos Excel con patrón {pattern} en {folder.resolve()}")
    return matches[0]  # el más reciente

def header_map(ws, header_row=1):
    """Devuelve dict: header_normalizado -> (col_idx, header_original)"""
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = str(v).strip().lower()
        hm[key] = (c, str(v))
    return hm

def ensure_col(ws, header_name, header_row=1):
    """Crea columna si no existe; retorna col_idx."""
    hm = header_map(ws, header_row)
    key = header_name.strip().lower()
    if key in hm:
        return hm[key][0]
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = header_name
    return new_col

def build_lookup_from_resumen(ws_resumen, doc_header="documento"):
    hm = header_map(ws_resumen)
    # Posibles nombres de columnas (ajusta si tus encabezados son distintos)
    doc_col = hm.get("documento", hm.get(doc_header.lower(), None))
    avance_col = hm.get("porcentaje de avance", hm.get("avance", None))
    matric_col = hm.get("número de matrículas", hm.get("numero de matriculas", hm.get("matriculas", None)))

    if not doc_col:
        raise ValueError("No encontré la columna 'Documento' en la hoja Resumen.")
    if not avance_col:
        raise ValueError("No encontré la columna 'Porcentaje de Avance' o 'avance' en la hoja Resumen.")
    if not matric_col:
        raise ValueError("No encontré la columna 'Número de Matrículas' o 'matriculas' en la hoja Resumen.")

    doc_col = doc_col[0]
    avance_col = avance_col[0]
    matric_col = matric_col[0]

    lookup = {}
    for r in range(2, ws_resumen.max_row + 1):
        doc = ws_resumen.cell(row=r, column=doc_col).value
        if doc is None:
            continue
        doc_key = str(doc).strip()
        avance = ws_resumen.cell(row=r, column=avance_col).value
        matric = ws_resumen.cell(row=r, column=matric_col).value
        lookup[doc_key] = (avance, matric)
    return lookup

def add_avance_matriculas_to_detalle(
    excel_path=None,
    folder=".",
    pattern="*.xlsx",
    sheet_resumen="Resumen",
    sheet_detalle="Detalle",
    doc_header_detalle="Documento"
):
    # 1) Encontrar archivo si no te pasan ruta
    if excel_path is None:
        excel_path = find_excel_file(folder=folder, pattern=pattern)

    wb = load_workbook(excel_path)
    if sheet_resumen not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_resumen}' en {excel_path}")
    if sheet_detalle not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_detalle}' en {excel_path}")

    ws_res = wb[sheet_resumen]
    ws_det = wb[sheet_detalle]

    # 2) Crear diccionario: Documento -> (avance, matriculas)
    lookup = build_lookup_from_resumen(ws_res)

    # 3) Asegurar columnas en Detalle
    det_headers = header_map(ws_det)
    doc_col_det = det_headers.get(doc_header_detalle.lower())
    if not doc_col_det:
        raise ValueError("No encontré la columna 'Documento' en la hoja Detalle.")
    doc_col_det = doc_col_det[0]

    col_avance = ensure_col(ws_det, "avance")
    col_matric = ensure_col(ws_det, "matriculas")

    # 4) Rellenar filas Detalle por Documento
    for r in range(2, ws_det.max_row + 1):
        doc = ws_det.cell(row=r, column=doc_col_det).value
        if doc is None:
            continue
        doc_key = str(doc).strip()
        if doc_key in lookup:
            avance, matric = lookup[doc_key]
            ws_det.cell(row=r, column=col_avance).value = avance
            ws_det.cell(row=r, column=col_matric).value = matric

    # 5) Guardar (sobrescribe o crea nuevo)
    out_path = Path(excel_path).with_name(Path(excel_path).stem + "_con_avance_matriculas.xlsx")
    wb.save(out_path)
    return str(out_path)

if __name__ == "__main__":
    salida = add_avance_matriculas_to_detalle(
        excel_path=None,         # o pon la ruta exacta: "mi_archivo.xlsx"
        folder=".",              # carpeta donde está el Excel
        pattern="Avances_puntaje.xlsx",        # patrón para buscar (puedes usar "reporte_*.xlsx")
        sheet_resumen="Resumen",
        sheet_detalle="Detalle",
        doc_header_detalle="Documento"
    )
    print("✅ Archivo generado:", salida)
